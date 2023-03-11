#代码by德浩
from click import style
import numpy as np
import openpyxl
from openpyxl import Workbook
#from examples.MDDQN.env import ArmEnv
import parl
from parl.utils import csv_logger, logger, ReplayMemory,summary
from model import Model
from cartpole_model import CartpoleModel
from cartpole_agent import CartpoleAgent
from parl.algorithms import DQN,DDQN
from env import ArmEnv

import matplotlib.pyplot as plt
import xlwt
#arduino模块
import serial
import time
import numpy as np
ser = serial.Serial('COM12',115200,timeout=1)

#训练的频率
LEARN_FREQ = 5  # training frequency
MEMORY_SIZE = 10000
MEMORY_WARMUP_SIZE = 100
BATCH_SIZE = 64
LEARNING_RATE = 0.0001
GAMMA = 0.95
action_memory=np.zeros(8)
total_step = 0
eval_step = 0
Gradient_Bandit_alpha=0.1


#保存数据训练的步数和每一步带来的奖励
wb1=xlwt.Workbook()
sheet1=wb1.add_sheet('train_steps&train_rewards')
text1 = ["total_tain_steps","reward_every_step"]
for i in range(len(text1)):
    sheet1.write(0,i,text1[i])
#保存每一步评估过程中的奖励
wb2=xlwt.Workbook()
sheet2=wb2.add_sheet('total_eval_steps&eval_rewards')
text2 = ["total_eval_steps","reward_every_step"]
for i in range(len(text2)):
    sheet2.write(0,i,text2[i])
#保存每一个episode中的平均奖励值，横坐标episode纵坐标train_平均奖励值
wb3=xlwt.Workbook()
sheet3=wb3.add_sheet('every_train_episdoe&rewards')
text3 = ["every_train_steps","reward_every_episode"]
for i in range(len(text3)):
    sheet3.write(0,i,text3[i])

#保存每个评估episode中的平均奖励值，横坐标eval——episode纵坐标eval_平均奖励值
wb4=xlwt.Workbook()
sheet4=wb4.add_sheet('every_eval_episdoe&rewards')
text4 = ["every_eval_episode","reward_every_eval_episode"]
for i in range(len(text4)):
    sheet4.write(0,i,text4[i])



# train an episode
def run_train_episode(agent, env, rpm):
    print("run_train_episode start...")
    total_reward = 0
    obs = env.reset()
    action_memory=np.zeros(8)
    step = 0
    global total_step
    #action_memory=np.zeros(8)
    while True:
        step += 1
        total_step+=1
        action = agent.sample(obs)
        print({"ACTION"},action)
        #action=1
        if action == 0:
            action_memory[0]+=1
        if action == 1:
            action_memory[1]+=1
        if action == 2:
            action_memory[2]+=1
        if action == 3:
            action_memory[3]+=1
        if action == 4:
            action_memory[4]+=1
        if action == 5:
            action_memory[5]+=1
        if action == 6:
            action_memory[6]+=1
        if action == 7:
            action_memory[7]+=1
        #向arduino发送电机驱动信号
        #ser.write(str(action).encode('utf-8'))
        ser.write(str(action).encode('utf-8'))
        #给arduino预留一些响应时间
        
        val2 = ser.readline().decode('utf-8')
        print({"VAL2"},val2)
        time.sleep(1.0)

        next_obs, reward, done, _ = env.step(action)
        sheet1.write(total_step,0,total_step)
        sheet1.write(total_step,1,reward)
        wb1.save('total_step&train_reward.xls')
        save_path = './model_TRAIN_1.ckpt'
        agent.save(save_path)
        #plt.show()
        rpm.append(obs, action, reward, next_obs, done)
        # train model
        if (len(rpm) > MEMORY_WARMUP_SIZE) and (step % LEARN_FREQ == 0):
            # s,a,r,s',done
            (batch_obs, batch_action, batch_reward, batch_next_obs,
             batch_done) = rpm.sample_batch(BATCH_SIZE)
            train_loss = agent.learn(batch_obs, batch_action, batch_reward,
                                     batch_next_obs, batch_done)

        total_reward += reward
        obs = next_obs
        if step == 50:
            done ==True
        if done:
            break
    return total_reward


# evaluate 5 episodes
def run_evaluate_episodes(agent, env, eval_episodes=1, render=False):
    print("run_evaluate_episodes")
    eval_reward = []
    global eval_step 
    evaluate_step = 0
    for i in range(eval_episodes):
        obs = env.reset()

        action_memory=np.zeros(8)
        episode_reward = 0
        print("here")
        while True:
            action = agent.predict(obs)
            #向arduino发送电机驱动信号
            #ser = serial.Serial('COM8',115200,timeout=1)
            if action == 0:
                action_memory[0]+=1
            if action == 1:
                action_memory[1]+=1
            if action == 2:
                action_memory[2]+=1
            if action == 3:
                action_memory[3]+=1
            if action == 4:
                action_memory[4]+=1
            if action == 5:
                action_memory[5]+=1
            if action == 6:
                action_memory[6]+=1
            if action == 7:
                action_memory[7]+=1            
            val = ser.write(str(action).encode('utf-8'))
            #给arduino预留一些响应时间
            time.sleep(1.0)
            #val2 = ser.readline().decode('utf-8')
            #print(val2)
            obs, reward, done, _ = env.step(action)
            episode_reward += reward
            eval_step+=1
            evaluate_step+=1

            #保存数据
            sheet2.write(eval_step,0,eval_step)
            sheet2.write(eval_step,1,reward)
            wb2.save('total_eval_step&reward.xls')

            if render:
                env.render()
            #if evaluate_step ==10:
            #    done = True
            if done:
                #get_back_to_original_pos()
                #ax5 = plt.subplot(324)
                #ax5.margins(0.05)
                #ax5.plot(eval_step,obs)
                #ax5.set_title(obs/eval_step)
                #ax6 = plt.subplot(325)
                #ax6.margins(0.05)
                #ax6.plot(eval_step,goal)
                #ax6.set_title(goal/eval_step)                
                break
        eval_reward.append(episode_reward)
    return np.mean(eval_reward)


def main():
    #env = gym.make('CartPole-v0')
    env = ArmEnv()
    #obs_dim = env.observation_space.shape[0]
    #状态量
    obs_dim = env.state_dim
    #动作维度
    act_dim = env.action_dim
    #act_dim = env.action_space.n
    logger.info('obs_dim {}, act_dim {}'.format(obs_dim, act_dim))

    # set action_shape = 0 while in discrete control environment
    rpm = ReplayMemory(MEMORY_SIZE, obs_dim, 0)

    # build an agent
    model = CartpoleModel(obs_dim=obs_dim, act_dim=act_dim)
    alg = DDQN(model, gamma=GAMMA, lr=LEARNING_RATE)
    agent = CartpoleAgent(
        alg, act_dim=act_dim, e_greed=0.30, e_greed_decrement=0.0001)
    agent.restore('./model_TRAIN_1.ckpt')
    eval_reward = run_evaluate_episodes(agent, env, render=False)
    #logger.info('episode:{}    e_greed:{}   Test reward:{}'.format(
    #    agent.e_greed, eval_reward))
        #保存数据


if __name__ == '__main__':
    main()
